#!/usr/bin/env python3
"""インスタ投稿スケジュールの「ブログ投稿日」列を、記事の実際の公開日に合わせる。

なぜ必要か:
  記事は下書きとして作ってからキューで公開されるため、xlsxに行を追加した時点の
  日付と、実際に公開された日付がズレる。publish.sh から自動で呼ばれるので、
  公開のたびにこの列が正しい日付に直る。

使い方:
  python3 sync-schedule.py            … 実際に直す
  python3 sync-schedule.py --check    … 直さずにズレだけ報告する
"""
import glob
import re
import sys
import unicodedata

import openpyxl

XLSX = "インスタ投稿スケジュール_ドタバタ母さんブログ.xlsx"
COL_NO, COL_TITLE, COL_DATE = 1, 2, 3


def normalize(s):
    """全角・半角や空白の違いを吸収してタイトルを突き合わせる"""
    return unicodedata.normalize("NFKC", str(s)).replace(" ", "")


def load_posts():
    """記事ファイルから タイトル → (公開日, 下書きかどうか) を作る"""
    posts = {}
    for path in glob.glob("content/posts/*.md"):
        try:
            front = open(path, encoding="utf-8").read().split("---")[1]
        except IndexError:
            continue
        title = re.search(r'^title: *"?(.*?)"?\s*$', front, re.M)
        date = re.search(r"^date: *(\S+)", front, re.M)
        draft = re.search(r"^draft: *(\S+)", front, re.M)
        if title and date:
            is_draft = (draft.group(1) if draft else "false") == "true"
            posts[normalize(title.group(1))] = (date.group(1).replace("-", "/"), is_draft)
    return posts


def main():
    check_only = "--check" in sys.argv
    posts = load_posts()
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    changed, unknown = [], []
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=COL_TITLE).value
        if not title:
            continue
        key = normalize(title)
        if key not in posts:
            unknown.append((ws.cell(row=row, column=COL_NO).value, str(title)[:34]))
            continue
        real_date, is_draft = posts[key]
        if is_draft:
            continue  # まだ公開していない記事は空欄のままにする
        cell = ws.cell(row=row, column=COL_DATE)
        current = str(cell.value).replace("-", "/").split(" ")[0] if cell.value else ""
        if current != real_date:
            changed.append((ws.cell(row=row, column=COL_NO).value, current or "(空)", real_date, str(title)[:34]))
            if not check_only:
                cell.value = real_date

    if changed and not check_only:
        wb.save(XLSX)

    label = "ズレている行" if check_only else "直した行"
    print(f"{label}: {len(changed)}件")
    for no, before, after, title in changed:
        print(f"  No.{no}  {before} → {after}   {title}")
    if unknown:
        print(f"\n記事が見つからない行: {len(unknown)}件（タイトルを変更した記事かもしれません）")
        for no, title in unknown[:10]:
            print(f"  No.{no}  {title}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
